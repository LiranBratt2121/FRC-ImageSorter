import cv2
from numpy import ndarray, random
import openpyxl
from openpyxl.drawing.image import Image
import os
import shutil

from typing import Dict, List
from dataclasses import dataclass

from utils import to_ROI


@dataclass
class Shirt:
    """
    A data structure representing a detected shirt.

    Attributes:
      team_name (str): Name of the team the shirt belongs to (without year and info).
      confidence (float): Confidence level of identifying the Shirt in the image (0.0 to 1.0).
      roi (ndarray): Region of interest (bounding box) for the member in the image.
      year (str | None): Year associated with the team (optional).
      other_data (str | None): Additional information about the team (optional).
    """
    team_name: str
    confidence: float
    roi: ndarray
    year: str = None
    other_data: str = None


class ImageData:
    """
    A data structure containing information about an image and detected shirts.

    Attributes:
      image_path (str): Path to the image.
      teams (Dict[str, List[Shirt]]): Dictionary mapping team names (without year and info) to a list of Shirt objects.
      team_occurrences (Dict[str, int]): Dictionary mapping team names (without year and info) to their occurrence count.

    """

    def __init__(self, image_path: str) -> None:
        self.image_path = image_path
        self.teams: Dict[str, List[Shirt]] = {}
        self.team_occurrences: Dict[str, int] = {}

    def __repr__(self) -> str:
        """
        Returns a string representation of the ImageData object.
        """
        return f'ImageData(image_path={self.image_path}, teams={self.teams}, team_occurrences={self.team_occurrences})'

    def add(self, team_name: str, confidence: float, roi: ndarray, year: str = None, other_data: str = None) -> None:
        """
        Adds a shirt to the image data with the specified information.

        Args:
        team_name (str): Name of the team the shirt belongs to (without year and info).
        confidence (float): Confidence level of identifying the shirt in the image (0.0 to 1.0).
        roi (ndarray): Region of interest (bounding box) for the shirt in the image.
        year (str, optional): Year associated with the team. Defaults to None.
        other_data (str, optional): Additional information about the shirt. Defaults to None.
        """
        if team_name not in self.teams:
            self.teams[team_name] = [Shirt(team_name, confidence, roi, year, other_data)]
            self.team_occurrences[team_name] = 1
            return

        self.teams[team_name].append(Shirt(team_name, confidence, roi, year, other_data))
        self.team_occurrences[team_name] += 1

    @property
    def get_team_names(self) -> List[str]:
        """
        Returns a list of unique team names (without year and additional info).

        Returns:
          List[str]: List of team names.
        """
        return list(self.teams.keys())

    @staticmethod
    def to_excel(image_data_list: List, output_filename: str) -> None:
        """
        Exports a list of ImageData objects to an Excel spreadsheet.

        Args:
            image_data_list (List[ImageData]): List of ImageData objects containing image data.
            output_filename (str): Name of the output Excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Image Data"
        temp_dir = 'temp'

        # Add temp directory for saved excel images
        try:
            os.mkdir(temp_dir)
        except FileExistsError:
            shutil.rmtree(temp_dir)
            os.mkdir(temp_dir)

        row_index = 1
        for image_data in image_data_list:
            # Load the full image
            full_image = cv2.imread(image_data.image_path)

            # Write image path
            ws.cell(row=row_index, column=1).value = "Image Path"
            ws.cell(row=row_index, column=2).value = image_data.image_path
            row_index += 1

            # Write full image
            full_image_name = f"full_image_{row_index}.png"
            cv2.imwrite(os.path.join(temp_dir, full_image_name),
                        cv2.resize(full_image, (256, 220)))
            ws.add_image(
                Image(os.path.join(temp_dir, full_image_name)), f"B{row_index}")
            row_index += 12  # Move to the next row after the full image

            for team_name, shirts in image_data.teams.items():
                for idx, shirt in enumerate(shirts):
                    # Write shirt data labels
                    ws.cell(row=row_index, column=1).value = "Team Name"
                    ws.cell(row=row_index + 1, column=1).value = "Confidence"
                    ws.cell(row=row_index + 2, column=1).value = "Year"
                    ws.cell(row=row_index + 3, column=1).value = "Other Data"

                    # Write shirt data values
                    ws.cell(row=row_index, column=2).value = team_name
                    ws.cell(row=row_index + 1,
                            column=2).value = shirt.confidence
                    ws.cell(row=row_index + 2,
                            column=2).value = shirt.year if shirt.year else "N/A"
                    ws.cell(
                        row=row_index + 3, column=2).value = shirt.other_data if shirt.other_data else "N/A"

                    # Extract and write ROIs
                    roi_image = to_ROI(full_image, shirt.roi)
                    roi_image_name = f"roi_image_{shirt.team_name}_{idx}_{random.rand()}.png"
                    cv2.imwrite(os.path.join(temp_dir, roi_image_name),
                                cv2.resize(roi_image, (128, 64)))
                    # Adjust column as needed
                    ws.add_image(
                        Image(os.path.join(temp_dir, roi_image_name)), f"C{row_index}")

                    # Move to the next shirt
                    row_index += 6  # Adjust based on the number of ROIs for each shirt

            row_index += 1  # Move to the next image

        wb.save(output_filename)
        shutil.rmtree(temp_dir, ignore_errors=True)
        print(f"Excel file saved to {output_filename}")

